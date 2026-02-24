
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from django.http import HttpResponse
from django.utils import timezone
from PIL import Image as PILImage
from io import BytesIO
import os

def generate_trip_excel(queryset, filename_prefix="trips_report"):
    """
    Generate a premium-styled Excel (.xlsx) report for Trip objects.
    Color-coded sections, alternating rows, conditional formatting, and IMAGES.
    """
    timestamp = timezone.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"{filename_prefix}_{timestamp}.xlsx"
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trips Report"
    ws.sheet_properties.tabColor = "4F81BD"

    # ═══════════════════════════════════════════════
    # STYLES
    # ═══════════════════════════════════════════════
    thin_border = Border(
        left=Side(style='thin', color='D0D5DD'),
        right=Side(style='thin', color='D0D5DD'),
        top=Side(style='thin', color='D0D5DD'),
        bottom=Side(style='thin', color='D0D5DD')
    )
    
    # Title row
    title_font = Font(name='Calibri', bold=True, size=16, color='1E3A5F')
    subtitle_font = Font(name='Calibri', size=11, color='667085')
    
    # Section header colors (group headers)
    SECTION_COLORS = {
        'trip':     PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid'),  # Dark Navy
        'people':   PatternFill(start_color='0D9488', end_color='0D9488', fill_type='solid'),   # Teal
        'time':     PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid'),   # Purple
        'checkin':  PatternFill(start_color='16A34A', end_color='16A34A', fill_type='solid'),   # Green
        'checkout': PatternFill(start_color='EA580C', end_color='EA580C', fill_type='solid'),   # Orange
        'location': PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid'),   # Blue
        'sub':      PatternFill(start_color='DB2777', end_color='DB2777', fill_type='solid'),   # Pink
    }
    
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Data row fills
    row_even = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    row_odd = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    # Conditional fills
    pass_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    fail_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    pass_font = Font(name='Calibri', bold=True, size=10, color='166534')
    fail_font = Font(name='Calibri', bold=True, size=10, color='DC2626')
    yes_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    yes_font = Font(name='Calibri', bold=True, size=10, color='92400E')
    data_font = Font(name='Calibri', size=10, color='344054')
    
    # ═══════════════════════════════════════════════
    # TITLE ROWS
    # ═══════════════════════════════════════════════
    ws.merge_cells('A1:AI1')  # Extended to AI for new columns
    title_cell = ws['A1']
    title_cell.value = '📋 Attendance Trips Report'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 36

    ws.merge_cells('A2:AI2')
    sub_cell = ws['A2']
    sub_cell.value = f'Generated: {timezone.now().strftime("%d %b %Y, %I:%M %p")}  •  Total Trips: {queryset.count()}'
    sub_cell.font = subtitle_font
    sub_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 22

    # ═══════════════════════════════════════════════
    # SECTION GROUP HEADER (Row 3)
    # ═══════════════════════════════════════════════
    # Updated ranges due to added image columns
    # Trip: A-E (5)
    # People: F-K (6) -> Added 2 image cols
    # Timing: L-N (3)
    # Checkin: O-U (7) -> Added 1 image col
    # Checkout: V-AB (7) -> Added 1 image col
    # Location: AC-AD (2)
    # Substitute: AE-AI (5) -> Added 3 image cols
    
    section_groups = [
        ('A3:E3',   '🚛 TRIP INFO',         'trip'),
        ('F3:K3',   '👥 PEOPLE',             'people'),
        ('L3:N3',   '⏰ TIMING',             'time'),
        ('O3:U3',   '✅ CHECK-IN VEHICLE',   'checkin'),
        ('V3:AB3',  '🔄 CHECK-OUT VEHICLE',  'checkout'),
        ('AC3:AD3', '📍 LOCATION',           'location'),
        ('AE3:AI3', '🔀 SUBSTITUTE',          'sub'),
    ]
    
    group_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    
    for cell_range, label, section_key in section_groups:
        ws.merge_cells(cell_range)
        cell = ws[cell_range.split(':')[0]]
        cell.value = label
        cell.font = group_font
        cell.fill = SECTION_COLORS[section_key]
        cell.alignment = center
    ws.row_dimensions[3].height = 28

    # ═══════════════════════════════════════════════
    # COLUMN HEADERS (Row 4)
    # ═══════════════════════════════════════════════
    headers_with_sections = [
        # Trip Info (A-E)
        ('S.No',           'trip'),
        ('Date',           'trip'),
        ('Status',         'trip'),
        ('Route',          'trip'),
        ('Ward',           'trip'),
        # People (F-K)
        ('Driver Name',    'people'),
        ('Driver ID',      'people'),
        ('Driver Photo',   'people'), # NEW
        ('Helper Name',    'people'),
        ('Helper ID',      'people'),
        ('Helper Photo',   'people'), # NEW
        # Timing (L-N)
        ('Check-In',       'time'),
        ('Check-Out',      'time'),
        ('Duration',       'time'),
        # Check-in Vehicle (O-U)
        ('Plate No.',      'checkin'),
        ('Device Photo',   'checkin'), # NEW
        ('Hooter',         'checkin'),
        ('Nagar Nigam',    'checkin'),
        ('Logo',           'checkin'),
        ('Result',         'checkin'),
        ('Failure Reason', 'checkin'),
        # Check-out Vehicle (V-AB)
        ('Plate No.',      'checkout'),
        ('Device Photo',   'checkout'), # NEW
        ('Hooter',         'checkout'),
        ('Nagar Nigam',    'checkout'),
        ('Logo',           'checkout'),
        ('Result',         'checkout'),
        ('Failure Reason', 'checkout'),
        # Location (AC-AD)
        ('Check-In GPS',   'location'),
        ('Check-Out GPS',  'location'),
        # Substitute (AE-AI)
        ('Sub Driver',     'sub'),
        ('Sub Name',       'sub'),
        ('Sub Phone',      'sub'),
        ('Sub Photo',      'sub'), # NEW
        ('Sub Helper Photo','sub'), # NEW
    ]

    # Sub-header tint
    SUB_HEADER_COLORS = {
        'trip':     PatternFill(start_color='2D4F7F', end_color='2D4F7F', fill_type='solid'),
        'people':   PatternFill(start_color='0F766E', end_color='0F766E', fill_type='solid'),
        'time':     PatternFill(start_color='6D28D9', end_color='6D28D9', fill_type='solid'),
        'checkin':  PatternFill(start_color='15803D', end_color='15803D', fill_type='solid'),
        'checkout': PatternFill(start_color='C2410C', end_color='C2410C', fill_type='solid'),
        'location': PatternFill(start_color='1D4ED8', end_color='1D4ED8', fill_type='solid'),
        'sub':      PatternFill(start_color='BE185D', end_color='BE185D', fill_type='solid'),
    }

    for col_num, (h_title, section) in enumerate(headers_with_sections, 1):
        cell = ws.cell(row=4, column=col_num, value=h_title)
        cell.font = header_font
        cell.fill = SUB_HEADER_COLORS[section]
        cell.alignment = center
        cell.border = thin_border
    ws.row_dimensions[4].height = 30
    
    ws.freeze_panes = 'A5'

    # ═══════════════════════════════════════════════
    # HELPER: Extract vehicle compliance details
    # ═══════════════════════════════════════════════
    def get_vehicle_details(vehicle_record):
        if not vehicle_record:
            return ('-', None, '-', '-', '-', 'No Scan', 'No Vehicle Scan')
        
        dets = vehicle_record.detections or {}
        has_hooter = '✅' if (dets.get('hooter') or dets.get('beacon') or dets.get('red_light')) else '❌'
        has_nn = '✅' if (dets.get('nagar_nigam') or dets.get('text_nn') or dets.get('nn_text')) else '❌'
        has_logo = '✅' if (dets.get('logo') or dets.get('swachh_bharat') or dets.get('emblem')) else '❌'
        
        plate = vehicle_record.plate_number or '—'
        result = '✅ Passed' if vehicle_record.compliance_passed else '❌ Failed'
        
        reasons = []
        if '❌' in has_hooter: reasons.append('Hooter')
        if '❌' in has_nn: reasons.append('Nagar Nigam')
        if '❌' in has_logo: reasons.append('Logo')
        if plate == '—': reasons.append('Plate')
        
        reason_str = ', '.join(reasons) if reasons else '—'
        return (plate, vehicle_record.vehicle_image, has_hooter, has_nn, has_logo, result, reason_str)

    # ═══════════════════════════════════════════════
    # HELPER: Embed Image in Cell
    # ═══════════════════════════════════════════════
    def embed_image(worksheet, row_idx, col_idx, image_field):
        """
        Resize and embed an image into a specific cell.
        Handles both local paths and S3-like storage if accessible via .open()
        """
        if not image_field:
            worksheet.cell(row=row_idx, column=col_idx, value="No Image").alignment = center
            return

        try:
            # Open image securely
            img = PILImage.open(image_field)
            
            # Convert to RGB (remove alpha channel if PNG to avoid errors with some excel viewers)
            if img.mode in ('RGBA', 'P'):
                img = img.convert('RGB')
            
            # Resize (Thumbnail)
            max_size = (80, 80)
            img.thumbnail(max_size, PILImage.LANCZOS)
            
            # Save to BytesIO to pass to Openpyxl
            img_byte_arr = BytesIO()
            img.save(img_byte_arr, format='JPEG')
            img_byte_arr.seek(0)
            
            # Create Openpyxl Image
            xl_img = OpenpyxlImage(img_byte_arr)
            
            # Create anchor (using cell coordinates)
            # Add a small padding
            col_letter = get_column_letter(col_idx)
            xl_img.anchor = f'{col_letter}{row_idx}'
            
            # Add to worksheet
            worksheet.add_image(xl_img)
            
            # Remove text value if any
            worksheet.cell(row=row_idx, column=col_idx, value="")
            
        except Exception as e:
            # Fallback to text if image fails
            # print(f"Image embed error: {e}")
            worksheet.cell(row=row_idx, column=col_idx, value="Err").alignment = center

    # ═══════════════════════════════════════════════
    # DATA ROWS (starting row 5)
    # ═══════════════════════════════════════════════
    for idx, trip in enumerate(queryset):
        row_num = idx + 5
        is_even = idx % 2 == 0
        row_fill = row_even if is_even else row_odd
        
        # Determine data for Driver
        if trip.is_substitute_driver:
            driver_name = f"🔄 SUBSTITUTE ({trip.substitute_driver_name})"
            driver_id = "SUBSTITUTE"
            img_driver = None  # Don't show original driver's login photo in main col
        else:
            driver_name = trip.driver.full_name if trip.driver else '—'
            driver_id = trip.driver.employee_id if trip.driver else '—'
            img_driver = trip.checkin_driver_detection.frame_image if trip.checkin_driver_detection else None

        # Determine data for Helper
        if trip.is_substitute_helper:
            helper_name = "🔄 SUBSTITUTE HELPER"
            helper_id = "SUBSTITUTE"
            img_helper = None
        else:
            helper_name = trip.helper.full_name if trip.helper else ('Skipped' if trip.helper_skipped else '—')
            helper_id = trip.helper.employee_id if trip.helper else '—'
            img_helper = trip.checkin_helper_detection.frame_image if trip.checkin_helper_detection else None
        
        ci_plate, ci_img, ci_hooter, ci_nn, ci_logo, ci_res, ci_reason = get_vehicle_details(trip.checkin_vehicle)
        co_plate, co_img, co_hooter, co_nn, co_logo, co_res, co_reason = get_vehicle_details(trip.checkout_vehicle)

        checkin_loc = f"{trip.checkin_latitude},{trip.checkin_longitude}" if trip.checkin_latitude else '—'
        checkout_loc = f"{trip.checkout_latitude},{trip.checkout_longitude}" if trip.checkout_latitude else '—'

        # Format duration nicely
        duration = '—'
        if trip.work_duration:
            total_secs = int(trip.work_duration.total_seconds())
            hours, remainder = divmod(total_secs, 3600)
            mins, _ = divmod(remainder, 60)
            duration = f"{hours}h {mins}m"

        # Status formatting
        status = trip.status.replace('_', ' ').title()

        # Prepare images (FieldFile objects)
        # img_driver/img_helper already set above logic
        
        img_sub_driver = trip.substitute_driver_photo if trip.is_substitute_driver else None
        img_sub_helper = trip.substitute_helper_photo if trip.is_substitute_helper else None

        # Row Text Data (use placeholders for image columns)
        row_data = [
            idx + 1,
            trip.date.strftime('%d-%b-%Y') if trip.date else '—',
            status,
            trip.route.name if trip.route else '—',
            trip.route.ward.name if trip.route and trip.route.ward else '—',
            
            driver_name,
            driver_id,
            '🔄 SUB' if trip.is_substitute_driver else '', # Driver Photo Text Placeholder
            
            helper_name,
            helper_id,
            '🔄 SUB' if trip.is_substitute_helper else '', # Helper Photo Text Placeholder
            
            trip.checkin_time.strftime('%I:%M %p') if trip.checkin_time else '—',
            trip.checkout_time.strftime('%I:%M %p') if trip.checkout_time else '—',
            duration,
            
            ci_plate,
            '', # Check-in Vehicle Photo
            ci_hooter, ci_nn, ci_logo, ci_res, ci_reason,
            
            co_plate,
            '', # Check-out Vehicle Photo
            co_hooter, co_nn, co_logo, co_res, co_reason,
            
            checkin_loc,
            checkout_loc,
            
            '🔄 Yes' if trip.is_substitute_driver else 'No',
            trip.substitute_driver_name or '—',
            trip.substitute_driver_phone or '—',
            '', # Sub Driver Photo
            '', # Sub Helper Photo
        ]
        
        # Write Text Data and Styles
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = center if col_num <= 5 or col_num >= 30 else left_align
            cell.fill = row_fill
        
        # Embed Images
        # Driver Photo (Col 8)
        embed_image(ws, row_num, 8, img_driver)
        # Helper Photo (Col 11)
        embed_image(ws, row_num, 11, img_helper)
        # Check-in Vehicle Photo (Col 16)
        embed_image(ws, row_num, 16, ci_img)
        # Check-out Vehicle Photo (Col 23)
        embed_image(ws, row_num, 23, co_img)
        # Sub Photo (Col 34)
        embed_image(ws, row_num, 34, img_sub_driver)
        # Sub Helper Photo (Col 35)
        embed_image(ws, row_num, 35, img_sub_helper)

        # Set Row Height for Images
        ws.row_dimensions[row_num].height = 65  # Sufficient for 80px thumbnail (approx 60pts)

        # GPS Hyperlinks (Check-In GPS=28, Check-Out GPS=29)
        link_font = Font(name='Calibri', size=10, color='0000FF', underline='single')
        
        # Check-in GPS (Col 28)
        if trip.checkin_latitude and trip.checkin_longitude:
            gps_cell = ws.cell(row=row_num, column=28)
            gps_cell.hyperlink = f"https://www.google.com/maps?q={trip.checkin_latitude},{trip.checkin_longitude}"
            gps_cell.font = link_font
        
        # Check-out GPS (Col 29)
        if trip.checkout_latitude and trip.checkout_longitude:
            gps_cell = ws.cell(row=row_num, column=29)
            gps_cell.hyperlink = f"https://www.google.com/maps?q={trip.checkout_latitude},{trip.checkout_longitude}"
            gps_cell.font = link_font

        # Result Logic (Cols: CI Res=20, CO Res=27)
        for result_col in [20, 27]:
            cell = ws.cell(row=row_num, column=result_col)
            if '✅' in str(cell.value):
                cell.fill = pass_fill
                cell.font = pass_font
                cell.fill = fail_fill
                cell.font = fail_font
        
        # Status Logic (Col 3)
        status_cell = ws.cell(row=row_num, column=3)
        if 'completed' in trip.status.lower():
            status_cell.fill = pass_fill
            status_cell.font = pass_font
        elif trip.status in ['incomplete', 'checkout_started']:
            status_cell.fill = fail_fill
            status_cell.font = fail_font
        else:
            status_cell.fill = yes_fill
            status_cell.font = yes_font
        
        # Substitute Highlight
        if trip.is_substitute_driver:
            for sub_col in range(30, 36): # Cols 30-35
                cell = ws.cell(row=row_num, column=sub_col)
                cell.fill = yes_fill
                cell.font = yes_font

    # ═══════════════════════════════════════════════
    # COLUMN WIDTHS
    # ═══════════════════════════════════════════════
    # Set explicit widths for image columns, auto for others
    # Image columns: 8, 11, 16, 23, 34, 35
    img_cols = [8, 11, 16, 23, 34, 35]
    for c in img_cols:
        col_letter = get_column_letter(c)
        ws.column_dimensions[col_letter].width = 12 # Approx 80px visual width

    min_widths = {
        1: 6,    # S.No
        2: 13,   # Date
        3: 15,   # Status
        4: 18,   # Route
        6: 18,   # Driver Name
        9: 18,   # Helper Name
        12: 12,  # CI Time
        13: 12,  # CO Time
    }
    
    for col_idx in range(1, len(headers_with_sections) + 1):
        if col_idx in img_cols:
            continue
            
        col_letter = get_column_letter(col_idx)
        max_len = min_widths.get(col_idx, 12)
        
        # Sample first 50 rows for width
        for row in ws.iter_rows(min_row=4, max_row=min(ws.max_row, 55), min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    val_len = len(str(cell.value))
                    max_len = max(max_len, min(val_len + 3, 30))
        
        ws.column_dimensions[col_letter].width = max_len

    # ═══════════════════════════════════════════════
    # SUMMARY ROW
    # ═══════════════════════════════════════════════
    total_trips = queryset.count()
    if total_trips > 0:
        summary_row = total_trips + 6
        ws.merge_cells(f'A{summary_row}:G{summary_row}')
        
        summary_cell = ws.cell(row=summary_row, column=1)
        completed = sum(1 for t in queryset if t.status == 'completed')
        subs = sum(1 for t in queryset if t.is_substitute_driver)
        summary_cell.value = f'📊 Total: {total_trips}  |  ✅ Completed: {completed}  |  🔄 Substitutes: {subs}'
        summary_cell.font = Font(name='Calibri', bold=True, size=12, color='1E3A5F')
        summary_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[summary_row].height = 30

    wb.save(response)
    return response
